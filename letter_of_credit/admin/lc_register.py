from django.contrib import admin
from letter_of_credit.models import LCRegister
from ajax_select import make_ajax_form
from ajax_select.admin import AjaxSelectAdmin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font
from datetime import datetime


class LCRegisterAdmin(AjaxSelectAdmin):
    form = make_ajax_form(
            LCRegister, {'ccy_obj': 'currency', }
    )

    list_display = ('mf', 'ba', 'lc_number', 'applicant', 'acct_numb', 'ccy_obj',
                    'lc_amt_org_ccy_fmt', 'estb_date', 'expiry_date', 'bene', 'advising_bank', 'mt_730_received_at',
                    'status', 'os_amount')

    list_display_links = ('mf', 'ba', 'lc_number',)

    search_fields = (
        'mf', 'ba', 'estb_date', 'lc_number', 'applicant', 'ccy_obj__code', 'acct_numb',
        'lc_amt_org_ccy', 'bene', 'advising_bank', 'mt_730_received_at','status', 'os_amount')

    actions = ('download_mt_730_not_received', 'download_selected')

    def lc_amt_org_ccy_fmt(self, obj):
        return '{:,.2f}'.format(obj.lc_amt_org_ccy)

    lc_amt_org_ccy_fmt.short_description = 'FX Amount'

    def download_selected(self, request, queryset):
        file_name = '%s.xlsx' % datetime.now().strftime('LC-REGISTER-%Y-%m-%d-%H-%S-%f')
        wb = Workbook()
        font = Font(bold=True)
        sheet = wb.active
        row = 1

        columns = (
            ('lc_number', 'LC Number'),
            ('estb_date', 'Estab. Date'),
            ('applicant', 'Applicant'),
            ('bene', 'Beneficiary Name'),
            ('ccy_obj', 'Currency', 'code',),
            ('lc_amt_org_ccy', 'FX Amount'),
            ('lc_amt_usd', 'LC Amount In USD'),
            ('acct_numb', 'Account Number'),
            ('description', 'Goods Description'),
            ('mf', 'Form M Number'),
            ('expiry_date', 'Expiry Date'),
            ('advising_bank', 'Advising Bank'),
        )

        sheet.cell(row=1, column=1, value='S/N').font = font

        for index, el in enumerate(columns):
            sheet.cell(row=1, column=index + 2, value=el[1].upper()).font = font

        row += 1
        row_index = 1

        for lc in queryset:
            for index, el in enumerate(columns):
                val = getattr(lc, el[0])
                if len(el) == 3:
                    val = getattr(val, el[2])

                sheet.cell(row=row, column=1, value=row_index)
                sheet.cell(row=row, column=index + 2, value=val)
            row += 1
            row_index += 1
        resp = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        resp['Content-Disposition'] = 'attachment; filename="%s"' % file_name
        return resp

    download_selected.short_description = 'Download selected lc'

    def download_mt_730_not_received(self, request, queryset):
        file_name = '%s.xlsx' % datetime.now().strftime('%Y-%m-%d-%H-%S-%f')
        wb = Workbook()
        sheet = wb.active
        row = 1
        row_index = 1

        sheet.cell(row=row, column=5, value='Date Estb.')
        row += 1

        for lc in LCRegister.objects.filter(mt_730_received_at__isnull=True):
            sheet.cell(row=row, column=1, value=row_index)
            sheet.cell(row=row, column=2, value=lc.applicant)
            sheet.cell(row=row, column=3, value=lc.lc_number)
            sheet.cell(row=row, column=4, value=lc.lc_amt_org_ccy)
            sheet.cell(row=row, column=5, value=lc.estb_date)  # .strftime('%d-%m-%Y')

            row += 1
            row_index += 1
        resp = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        resp['Content-Disposition'] = 'attachment; filename="%s"' % file_name
        return resp

    download_mt_730_not_received.short_description = 'Download MT730 not received'


admin.site.register(LCRegister, LCRegisterAdmin)
