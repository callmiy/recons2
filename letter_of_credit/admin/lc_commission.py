from datetime import datetime

from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook

from letter_of_credit.models import LcCommission


@admin.register(LcCommission)
class LcCommissionAdmin(admin.ModelAdmin):
    list_display = (
        'lc_number', 'applicant', 'charge_date', 'ccy_code', 'transaction_amount_formatted', 'charge_amount_formatted',
        'percent_applied', 'exchange_rate', 'acct_numb', 'event', 'created_at',)

    search_fields = ('lc__lc_number', 'transaction_amount', 'charge_amount', 'exchange_rate', 'percent_applied',
                     'acct_numb', 'event', 'comment', 'charge_date', 'created_at', 'lc__applicant', 'lc__ccy_obj__code',
                     'lc__mf',)

    actions = ('download_selected',)

    def download_selected(self, request, queryset):
        file_name = '%s.xlsx' % datetime.now().strftime('LC-COMMISSION-%Y-%m-%d-%H-%S-%f')
        wb = Workbook()
        font = Font(bold=True)
        sheet = wb.active
        row = 1

        columns = (
            ('lc_number', 'LC Number',),
            ('applicant', 'Applicant',),
            ('charge_date', 'Charge Date'),
            ('ccy_code', 'Ccy',),
            ('lc_value', 'LC Value',),
            ('transaction_amount', 'Transaction Amount',),
            ('charge_amount', 'Charge Amount'),
            ('percent_applied', '% Applied'),
            ('exchange_rate', 'X Rate'),
            ('acct_numb', 'Account'),
        )

        sheet.cell(row=1, column=1, value='S/N').font = font

        for index, el in enumerate(columns):
            sheet.cell(row=1, column=index + 2, value=el[1].upper()).font = font

        row += 1
        row_index = 1

        for lc in queryset:
            for index, el in enumerate(columns):
                val = getattr(lc, el[0])
                sheet.cell(row=row, column=1, value=row_index)
                sheet.cell(row=row, column=index + 2, value=(callable(val) and val() or val))
            row += 1
            row_index += 1
        resp = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        resp['Content-Disposition'] = 'attachment; filename="%s"' % file_name
        return resp
