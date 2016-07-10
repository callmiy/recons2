from datetime import datetime, date

from django.http import HttpResponse
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.writer.excel import save_virtual_workbook

from core_recons.utilities import col
from letter_of_credit.models import ConsolidatedLcBidRequest


class DownloadConsolidatedBidsView(View):
    def set_header_row(self, sheet):
        font = Font(bold=True, name='Rockwell')
        alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        s = sheet.cell(row=1, column=col.a, value='S/N')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.b, value='DATE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.c, value='LC/BC REF')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.d, value='FORM M NUMBER')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.e, value='RATE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.f, value='CUSTOMER NAME')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.g, value='CURRENCY')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.h, value='FORM M VALUE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.i, value='ACTUAL AMOUNT FOR BIDDING')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.j, value='TOTAL ALLOCATION SO FAR')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.l, value='OUTSTANDING BALANCE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.m, value='ACCOUNT NOS')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.n, value='ITEM OF IMPORT')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.o, value='MATURITY DATE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=col.p, value='CATEGORY')
        s.font = font
        s.alignment = alignment

    def get(self, request):
        wb = Workbook()
        file_name = '%s.xlsx' % datetime.now().strftime('CONSOLIDATED-FOREX-DEMAND-%Y-%m-%d-%H-%S-%f')
        bid_ids = ConsolidatedLcBidRequest.objects.all()

        if bid_ids.exists():
            sheet = wb.active
            sheet.title = 'CASH BACKED'
            self.set_header_row(sheet)
            row = 2
            row_index = 1

            for bid in bid_ids:
                mf = bid.mf
                sheet.cell(row=row, column=col.a, value=row_index)
                sheet.cell(row=row, column=col.b, value=bid.created_at)

                lc_number = 'NEW LC'
                lc = mf.lc

                if lc:
                    lc_number = lc.lc_number

                sheet.cell(row=row, column=col.c, value=lc_number)
                sheet.cell(row=row, column=col.d, value=mf.number)
                sheet.cell(row=row, column=col.e, value=bid.rate)
                applicant = mf.applicant
                sheet.cell(row=row, column=col.f, value=applicant.name)
                sheet.cell(row=row, column=col.g, value=mf.currency.code)
                sheet.cell(row=row, column=col.h, value=mf.amount)
                sheet.cell(row=row, column=col.i, value=bid.sum_bid_requests())
                sheet.cell(row=row, column=col.j, value=bid.sum_allocations())
                sheet.cell(row=row, column=col.l, value=bid.outstanding_amount())

                acct = ''
                acct_numbers_qs = applicant.acct_numbers

                if acct_numbers_qs:
                    acct = acct_numbers_qs[0].nuban

                sheet.cell(row=row, column=col.m, value=acct)
                sheet.cell(row=row, column=col.n, value=mf.goods_description)
                sheet.cell(row=row, column=col.o, value='CASH BACKED')
                sheet.cell(row=row, column=col.p, value=bid.goods_category)

                row += 1
                row_index += 1

        resp = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        resp['Content-Disposition'] = 'attachment; filename="%s"' % file_name
        return resp
