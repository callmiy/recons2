from datetime import datetime, date

from django.http import HttpResponse
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.writer.excel import save_virtual_workbook

from letter_of_credit.models import LcBidRequest


class DownloadBidsView(View):
    def set_header_row(self, sheet, mark_as_downloaded):
        font = Font(bold=True, name='Rockwell')
        alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        header_col_1 = sheet.cell(row=1, column=1, value='DATE')
        header_col_1.font = font
        header_col_1.alignment = alignment

        s = sheet.cell(row=1, column=2, value='FORM M NUMBER')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=3, value='LC/BC REF')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=4, value='CUSTOMER NAME')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=5, value='CURRENCY')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=6, value='LC VALUE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=7, value='ACTUAL BID AMOUNT')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=8, value='FX ALLOCATION')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=9, value='OUTSTANDING BALANCE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=10, value='ACCOUNT NOS')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=11, value='ITEM OF IMPORT')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=12, value='MATURITY DATE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=13, value='CATEGORY')
        s.font = font
        s.alignment = alignment

        if not mark_as_downloaded:
            # we are downloading all bids
            header_col_1.value = 'S/N'
            s = sheet.cell(row=1, column=14, value='DATE SENT TO TREASURY')
            s.font = font
            s.alignment = alignment

            s = sheet.cell(row=1, column=15, value='REMARK')
            s.font = font
            s.alignment = alignment

    def get(self, request):
        wb = Workbook()
        bid_ids = request.GET.getlist('bid_ids')
        mark_as_downloaded = False

        if bid_ids:
            file_name = '%s.xlsx' % datetime.now().strftime('fx-request-%Y-%m-%d-%H-%S-%f')
            mark_as_downloaded = True
        else:
            file_name = '%s.xlsx' % datetime.now().strftime('all-bids-%Y-%m-%d-%H-%S-%f')
            bid_ids = LcBidRequest.objects.values_list('pk', flat=True)

        if bid_ids:
            sheet = wb.active
            self.set_header_row(sheet, mark_as_downloaded)
            row = 2
            row_index = 1

            for bid in LcBidRequest.objects.filter(pk__in=bid_ids):
                mf = bid.mf
                applicant = mf.applicant
                sheet.cell(
                        row=row, column=1, value=mark_as_downloaded and date.today().strftime('%d-%b-%Y') or row_index)
                sheet.cell(row=row, column=2, value=mf.number)

                lc_number = 'NEW LC'
                lc = mf.lc

                if lc:
                    lc_number = lc.lc_number

                sheet.cell(row=row, column=3, value=lc_number)
                sheet.cell(row=row, column=4, value=applicant.name)
                sheet.cell(row=row, column=5, value=mf.currency.code)

                total_allocation = sum([allocation['amount_allocated'] for allocation in bid.allocations()])
                outstanding_bid_amount = bid.amount - total_allocation
                sheet.cell(row=row, column=6, value=mf.amount)
                sheet.cell(row=row, column=7, value=outstanding_bid_amount)
                sheet.cell(row=row, column=8, value=total_allocation)
                sheet.cell(row=row, column=9, value=outstanding_bid_amount)

                acct = ''
                acct_numbers_qs = applicant.acct_numbers

                if acct_numbers_qs:
                    acct = acct_numbers_qs[0].nuban

                sheet.cell(row=row, column=10, value=acct)
                sheet.cell(row=row, column=11, value=mf.goods_description)
                sheet.cell(row=row, column=12, value='CASH BACKED')

                if not mark_as_downloaded:
                    remark = ''

                    if bid.deleted_at:
                        remark = 'bid deleted'
                    elif mf.deleted_at:
                        remark = 'form M cancelled'

                    sheet.cell(row=row, column=14, value=bid.requested_at)
                    sheet.cell(row=row, column=15, value=remark)

                row += 1
                row_index += 1

                if mark_as_downloaded and not bid.downloaded:
                    bid.downloaded = True
                    bid.save()

        resp = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        resp['Content-Disposition'] = 'attachment; filename="%s"' % file_name
        return resp


class DownloadBidsLcEstablished(View):
    def set_header_row(self, sheet):
        font = Font(bold=True, name='Rockwell')
        alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        header_col_1 = sheet.cell(row=1, column=1, value='S/N')
        header_col_1.font = font
        header_col_1.alignment = alignment

        s = sheet.cell(row=1, column=2, value='APPLICANT')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=3, value='ESTB. DATE')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=4, value='LC NUMBER')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=5, value='MF NUMBER')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=6, value='CCY')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=7, value='LC AMOUNT')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=8, value='BID AMOUNT')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=9, value='ALLOCATIONS')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=10, value='OUTSTANDING BIDS')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=11, value='ADVISING BANK')
        s.font = font
        s.alignment = alignment

        s = sheet.cell(row=1, column=12, value='COMMENT')
        s.font = font
        s.alignment = alignment

    def get(self, request):
        file_name = '%s.xlsx' % datetime.now().strftime('bids-lc-established-%Y-%m-%d-%H-%S-%f')
        wb = Workbook()
        sheet = wb.active
        self.set_header_row(sheet, )
        row = 1
        row_index = 0

        for bid in LcBidRequest.objects.filter(mf__lc__isnull=False):
            remark = ''

            if bid.deleted_at:
                remark = 'bid deleted'
            elif bid.mf.deleted_at:
                remark = 'form M cancelled'

            row_index += 1
            row += 1
            sheet.cell(row=row, column=1, value=row_index)
            lc = bid.mf.lc
            allocations = sum([allocation['amount_allocated'] for allocation in bid.allocations()])
            sheet.cell(row=row, column=2, value=bid.applicant())
            sheet.cell(row=row, column=3, value=lc.lc_number)
            sheet.cell(row=row, column=4, value=lc.estb_date)
            sheet.cell(row=row, column=5, value=bid.form_m_number())
            sheet.cell(row=row, column=6, value=bid.currency().code)
            sheet.cell(row=row, column=7, value=lc.lc_amt_org_ccy)
            sheet.cell(row=row, column=8, value=bid.amount)
            sheet.cell(row=row, column=9, value=allocations)
            sheet.cell(row=row, column=10, value=(bid.amount - allocations))
            sheet.cell(row=row, column=11, value=lc.advising_bank)
            sheet.cell(row=row, column=12, value=remark)

        resp = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        resp['Content-Disposition'] = 'attachment; filename="%s"' % file_name
        return resp
