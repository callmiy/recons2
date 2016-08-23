from datetime import datetime
from django.http import HttpResponse
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.writer.excel import save_virtual_workbook

from letter_of_credit.models import TreasuryAllocation

border = Border(top=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                left=Side(border_style='thin', color='FF000000'))


class DownloadAllocationsView(View):
    def set_header_row(self, sheet):
        font = Font(bold=True, name='Rockwell')
        alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        s = sheet.cell(row=1, column=1, value='S/N')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['A'].width = 5

        s = sheet.cell(row=1, column=2, value='DEAL NO.')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['B'].width = 11.9

        s = sheet.cell(row=1, column=3, value='SOURCE')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['C'].width = 10

        s = sheet.cell(row=1, column=4, value='DEAL DATE')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['D'].width = 10.5

        s = sheet.cell(row=1, column=5, value='REF')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['E'].width = 14.8

        s = sheet.cell(row=1, column=6, value='NAME')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['F'].width = 23

        s = sheet.cell(row=1, column=7, value='CCY')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['G'].width = 6

        s = sheet.cell(row=1, column=8, value='AMOUNT')
        s.font = font
        s.alignment = alignment
        s.border = border
        sheet.column_dimensions['H'].width = 12.3

        s = sheet.cell(row=1, column=9, value='RATE')
        s.font = font
        s.border = border
        s.alignment = alignment

        s = sheet.cell(row=1, column=10, value='CLIENT CATEGORY')
        s.font = font
        s.border = border
        s.alignment = alignment

        s = sheet.cell(row=1, column=11, value='PRODUCT TYPE')
        s.font = font
        s.border = border
        s.alignment = alignment

        s = sheet.cell(row=1, column=12, value='TRANSACTION TYPE')
        s.font = font
        s.border = border
        s.alignment = alignment

    def get(self, request):
        wb = Workbook()
        file_name = '%s.xlsx' % datetime.now().strftime('fx-deals-%Y-%m-%d-%H-%S-%f')
        allocation_ids = request.GET.getlist('allocation_ids')

        if allocation_ids:
            qs = TreasuryAllocation.objects.filter(pk__in=allocation_ids)
        else:
            qs = TreasuryAllocation.objects.all()

        sheet = wb.active
        self.set_header_row(sheet)
        row = 2
        row_index = 1

        for allocation in qs:
            c = sheet.cell(row=row, column=1, value=row_index)
            c.border = border

            c = sheet.cell(row=row, column=2, value=allocation.deal_number)
            c.border = c.border = border

            c = sheet.cell(row=row, column=3, value=allocation.source_of_fund)
            c.border = border

            c = sheet.cell(row=row, column=4, value=allocation.deal_date)
            c.number_format = 'dd/mm/yyyy'
            c.border = border

            c = sheet.cell(row=row, column=5, value=allocation.ref)
            c.border = border

            c = sheet.cell(row=row, column=6, value=allocation.customer_name_no_ref)
            c.border = border

            c = sheet.cell(row=row, column=7, value=allocation.currency)
            c.border = border

            c = sheet.cell(row=row, column=8, value=(-1 * allocation.fcy_amount))
            c.number_format = FORMAT_NUMBER_COMMA_SEPARATED1
            c.border = border

            c = sheet.cell(row=row, column=9, value=allocation.naira_rate)
            c.number_format = '#,##0.0000'
            c.border = border

            c = sheet.cell(row=row, column=10, value=allocation.client_category)
            c.border = border

            c = sheet.cell(row=row, column=11, value=allocation.product_type)
            c.border = border

            c = sheet.cell(row=row, column=12, value=allocation.transaction_type)
            c.border = border

            row += 1
            row_index += 1

        resp = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        resp['Content-Disposition'] = 'attachment; filename="%s"' % file_name
        return resp
